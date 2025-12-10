"""
Historical Import Service

Service for importing historical actuals from Excel/CSV files.
Supports multiple modules (enrollment, revenue, costs, dhg, capex).
"""

from __future__ import annotations

import csv
import io
import uuid
from decimal import Decimal, InvalidOperation
from enum import Enum
from typing import Any

from openpyxl import load_workbook
from pydantic import BaseModel
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.analysis import HistoricalActuals


class ImportModule(str, Enum):
    """Supported modules for historical import."""

    ENROLLMENT = "enrollment"
    DHG = "dhg"
    REVENUE = "revenue"
    COSTS = "costs"
    CAPEX = "capex"


class ImportStatus(str, Enum):
    """Status of an import record."""

    VALID = "valid"
    WARNING = "warning"
    ERROR = "error"


class ImportRecordResult(BaseModel):
    """Result for a single import record."""

    row_number: int
    status: ImportStatus
    module: str
    dimension_type: str
    dimension_code: str | None
    value: Decimal | int | None
    message: str | None = None


class ImportPreviewResult(BaseModel):
    """Result of previewing an import file."""

    fiscal_year: int
    detected_module: str | None
    total_rows: int
    valid_rows: int
    invalid_rows: int
    warnings: list[str]
    errors: list[str]
    sample_data: list[dict[str, str | int | float | None]]
    can_import: bool


class ImportResultStatus(str, Enum):
    """Status of import result."""

    SUCCESS = "success"
    PARTIAL = "partial"
    ERROR = "error"


class ImportResult(BaseModel):
    """Result of importing historical data."""

    fiscal_year: int
    module: str
    status: ImportResultStatus
    imported_count: int
    updated_count: int
    skipped_count: int
    error_count: int
    message: str
    errors: list[str]


# Column mappings for different modules
ENROLLMENT_COLUMNS = {
    "fiscal_year": ["fiscal_year", "year", "fy"],
    "level_code": ["level_code", "level", "code"],
    "level_name": ["level_name", "name"],
    "student_count": ["student_count", "students", "count", "enrollment"],
}

FINANCIAL_COLUMNS = {
    "fiscal_year": ["fiscal_year", "year", "fy"],
    "account_code": ["account_code", "account", "code"],
    "account_name": ["account_name", "name", "description"],
    "annual_amount": ["annual_amount", "amount", "annual_amount_sar", "amount_sar"],
}

DHG_COLUMNS = {
    "fiscal_year": ["fiscal_year", "year", "fy"],
    "subject_code": ["subject_code", "subject", "code"],
    "subject_name": ["subject_name", "name"],
    "fte_count": ["fte_count", "fte", "count"],
    "hours": ["hours", "total_hours", "annual_hours"],
}


class HistoricalImportService:
    """Service for importing historical actuals."""

    def __init__(self, session: AsyncSession) -> None:
        self.session = session

    async def preview_import(
        self,
        file_content: bytes,
        filename: str,
        fiscal_year: int,
        module: ImportModule | None = None,
    ) -> ImportPreviewResult:
        """
        Preview an import file without saving.

        Args:
            file_content: Raw file bytes
            filename: Original filename (used to detect format)
            fiscal_year: Target fiscal year
            module: Optional module filter

        Returns:
            ImportPreviewResult with validation details
        """
        # Parse file based on extension
        if filename.endswith(".csv"):
            rows = self._parse_csv(file_content)
        elif filename.endswith((".xlsx", ".xls")):
            rows = self._parse_excel(file_content)
        else:
            return ImportPreviewResult(
                fiscal_year=fiscal_year,
                detected_module=None,
                total_rows=0,
                valid_rows=0,
                invalid_rows=0,
                warnings=[],
                errors=["Unsupported file format. Use .xlsx or .csv"],
                sample_data=[],
                can_import=False,
            )

        # Detect module from columns if not specified
        detected_module = module or self._detect_module(rows)

        # Validate records
        results: list[ImportRecordResult] = []
        warnings: list[str] = []
        errors: list[str] = []

        for idx, row in enumerate(rows, start=2):  # Start at 2 (header is row 1)
            result = self._validate_record(row, idx, fiscal_year, detected_module)
            results.append(result)

            # Collect warnings and errors with row context
            if result.status == ImportStatus.WARNING and result.message:
                warnings.append(f"Row {result.row_number}: {result.message}")
            elif result.status == ImportStatus.ERROR and result.message:
                errors.append(f"Row {result.row_number}: {result.message}")

        # Calculate totals
        valid_count = sum(1 for r in results if r.status == ImportStatus.VALID)
        error_count = sum(1 for r in results if r.status == ImportStatus.ERROR)

        # Convert first few rows to sample_data format
        sample_data: list[dict[str, str | int | float | None]] = []
        for row in rows[:5]:  # First 5 rows as sample
            sample_row: dict[str, str | int | float | None] = {}
            for key, value in row.items():
                # Try to convert numeric strings to numbers
                if value:
                    try:
                        # Try int first
                        sample_row[key] = int(value)
                    except ValueError:
                        try:
                            # Then try float
                            sample_row[key] = float(value)
                        except ValueError:
                            # Keep as string
                            sample_row[key] = value
                else:
                    sample_row[key] = None
            sample_data.append(sample_row)

        return ImportPreviewResult(
            fiscal_year=fiscal_year,
            detected_module=detected_module.value if detected_module else None,
            total_rows=len(results),
            valid_rows=valid_count,
            invalid_rows=error_count,
            warnings=warnings[:10],  # Limit to first 10 warnings
            errors=errors[:10],  # Limit to first 10 errors
            sample_data=sample_data,
            can_import=error_count == 0 and valid_count > 0,
        )

    async def import_data(
        self,
        file_content: bytes,
        filename: str,
        fiscal_year: int,
        module: ImportModule | None = None,
        overwrite: bool = False,
        user_id: uuid.UUID | None = None,
    ) -> ImportResult:
        """
        Import historical data from file.

        Args:
            file_content: Raw file bytes
            filename: Original filename
            fiscal_year: Target fiscal year
            module: Optional module filter
            overwrite: Whether to replace existing data
            user_id: User performing the import

        Returns:
            ImportResult with import status
        """
        # Parse file
        if filename.endswith(".csv"):
            rows = self._parse_csv(file_content)
        elif filename.endswith((".xlsx", ".xls")):
            rows = self._parse_excel(file_content)
        else:
            return ImportResult(
                fiscal_year=fiscal_year,
                module=module.value if module else "unknown",
                status=ImportResultStatus.ERROR,
                imported_count=0,
                updated_count=0,
                skipped_count=0,
                error_count=0,
                message="Unsupported file format. Use .xlsx or .csv",
                errors=["Unsupported file format"],
            )

        if not rows:
            return ImportResult(
                fiscal_year=fiscal_year,
                module=module.value if module else "unknown",
                status=ImportResultStatus.ERROR,
                imported_count=0,
                updated_count=0,
                skipped_count=0,
                error_count=0,
                message="No data found in file",
                errors=["Empty file"],
            )

        # Detect module
        detected_module = module or self._detect_module(rows)

        # Delete existing data if overwrite (count as updates)
        updated_count = 0
        if overwrite:
            updated_count = await self._delete_existing(fiscal_year, detected_module)

        # Import records
        imported = 0
        skipped = 0
        errors: list[str] = []

        for idx, row in enumerate(rows, start=2):
            try:
                record = self._create_record(row, fiscal_year, detected_module)
                if record:
                    self.session.add(record)
                    imported += 1
                else:
                    skipped += 1
            except Exception as e:
                errors.append(f"Row {idx}: {e!s}")
                skipped += 1

        # Commit transaction
        try:
            await self.session.commit()
        except Exception as e:
            await self.session.rollback()
            return ImportResult(
                fiscal_year=fiscal_year,
                module=detected_module.value,
                status=ImportResultStatus.ERROR,
                imported_count=0,
                updated_count=0,
                skipped_count=len(rows),
                error_count=1,
                message=f"Database error: {e!s}",
                errors=[str(e)],
            )

        # Determine status
        if imported > 0 and len(errors) == 0:
            status = ImportResultStatus.SUCCESS
        elif imported > 0 and len(errors) > 0:
            status = ImportResultStatus.PARTIAL
        else:
            status = ImportResultStatus.ERROR

        return ImportResult(
            fiscal_year=fiscal_year,
            module=detected_module.value,
            status=status,
            imported_count=imported,
            updated_count=updated_count,
            skipped_count=skipped,
            error_count=len(errors),
            message=f"Successfully imported {imported} records",
            errors=errors[:10],  # Limit errors to first 10
        )

    async def get_import_history(
        self,
        fiscal_year: int | None = None,
        module: ImportModule | None = None,
        limit: int = 50,
    ) -> list[dict[str, Any]]:
        """
        Get summary of imported historical data.

        Args:
            fiscal_year: Optional filter by year
            module: Optional filter by module
            limit: Maximum records to return

        Returns:
            List of import summaries
        """
        from sqlalchemy import func

        query = select(
            HistoricalActuals.fiscal_year,
            HistoricalActuals.module_code,
            func.count().label("record_count"),
            func.max(HistoricalActuals.created_at).label("last_imported"),
        ).group_by(
            HistoricalActuals.fiscal_year,
            HistoricalActuals.module_code,
        )

        if fiscal_year:
            query = query.where(HistoricalActuals.fiscal_year == fiscal_year)
        if module:
            query = query.where(HistoricalActuals.module_code == module.value)

        query = query.order_by(
            HistoricalActuals.fiscal_year.desc(),
            HistoricalActuals.module_code,
        ).limit(limit)

        result = await self.session.execute(query)
        rows = result.all()

        return [
            {
                "fiscal_year": row.fiscal_year,
                "module": row.module_code,
                "record_count": row.record_count,
                "last_imported": row.last_imported.isoformat() if row.last_imported else None,
            }
            for row in rows
        ]

    def _parse_csv(self, content: bytes) -> list[dict[str, str]]:
        """Parse CSV file content."""
        text = content.decode("utf-8-sig")  # Handle BOM
        reader = csv.DictReader(io.StringIO(text))
        return [row for row in reader]

    def _parse_excel(self, content: bytes) -> list[dict[str, str]]:
        """Parse Excel file content."""
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active

        if not sheet:
            return []

        rows: list[dict[str, str]] = []
        headers: list[str] = []

        for idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if idx == 0:
                # First row is headers
                headers = [str(cell).lower().strip() if cell else f"col_{i}" for i, cell in enumerate(row)]
            else:
                # Data rows
                row_dict = {
                    headers[i]: str(cell) if cell is not None else ""
                    for i, cell in enumerate(row)
                    if i < len(headers)
                }
                if any(row_dict.values()):  # Skip empty rows
                    rows.append(row_dict)

        return rows

    def _detect_module(self, rows: list[dict[str, str]]) -> ImportModule:
        """Detect module type from column names."""
        if not rows:
            return ImportModule.ENROLLMENT

        columns = set(rows[0].keys())
        lower_columns = {c.lower() for c in columns}

        # Check for enrollment columns
        if any(c in lower_columns for c in ["level_code", "level", "student_count", "students"]):
            return ImportModule.ENROLLMENT

        # Check for DHG columns
        if any(c in lower_columns for c in ["subject_code", "subject", "fte_count", "fte"]):
            return ImportModule.DHG

        # Check for financial columns (revenue/costs)
        if any(c in lower_columns for c in ["account_code", "account"]):
            # Determine if revenue or costs based on account code prefix
            for row in rows[:5]:
                account = self._get_column_value(row, FINANCIAL_COLUMNS["account_code"])
                if account and account.startswith("7"):
                    return ImportModule.REVENUE
                if account and account.startswith("6"):
                    return ImportModule.COSTS

            return ImportModule.REVENUE  # Default to revenue

        return ImportModule.ENROLLMENT

    def _get_column_value(self, row: dict[str, str], possible_names: list[str]) -> str | None:
        """Get value from row using possible column names."""
        for name in possible_names:
            for key in row.keys():
                if key.lower().strip() == name.lower():
                    return row[key].strip() if row[key] else None
        return None

    def _validate_record(
        self,
        row: dict[str, str],
        row_number: int,
        fiscal_year: int,
        module: ImportModule,
    ) -> ImportRecordResult:
        """Validate a single import record."""
        if module == ImportModule.ENROLLMENT:
            return self._validate_enrollment_record(row, row_number, fiscal_year)
        elif module == ImportModule.DHG:
            return self._validate_dhg_record(row, row_number, fiscal_year)
        else:
            return self._validate_financial_record(row, row_number, fiscal_year, module)

    def _validate_enrollment_record(
        self,
        row: dict[str, str],
        row_number: int,
        fiscal_year: int,
    ) -> ImportRecordResult:
        """Validate enrollment record."""
        level_code = self._get_column_value(row, ENROLLMENT_COLUMNS["level_code"])
        count_str = self._get_column_value(row, ENROLLMENT_COLUMNS["student_count"])

        if not level_code:
            return ImportRecordResult(
                row_number=row_number,
                status=ImportStatus.ERROR,
                module="enrollment",
                dimension_type="level",
                dimension_code=None,
                value=None,
                message="Missing level_code",
            )

        try:
            count = int(count_str) if count_str else 0
            if count < 0:
                raise ValueError("Negative count")
        except (ValueError, TypeError):
            return ImportRecordResult(
                row_number=row_number,
                status=ImportStatus.ERROR,
                module="enrollment",
                dimension_type="level",
                dimension_code=level_code,
                value=None,
                message=f"Invalid student count: {count_str}",
            )

        return ImportRecordResult(
            row_number=row_number,
            status=ImportStatus.VALID,
            module="enrollment",
            dimension_type="level",
            dimension_code=level_code,
            value=count,
        )

    def _validate_dhg_record(
        self,
        row: dict[str, str],
        row_number: int,
        fiscal_year: int,
    ) -> ImportRecordResult:
        """Validate DHG record."""
        subject_code = self._get_column_value(row, DHG_COLUMNS["subject_code"])
        fte_str = self._get_column_value(row, DHG_COLUMNS["fte_count"])

        if not subject_code:
            return ImportRecordResult(
                row_number=row_number,
                status=ImportStatus.ERROR,
                module="dhg",
                dimension_type="subject",
                dimension_code=None,
                value=None,
                message="Missing subject_code",
            )

        try:
            fte = Decimal(fte_str) if fte_str else Decimal("0")
            if fte < 0:
                raise ValueError("Negative FTE")
        except (ValueError, TypeError, InvalidOperation):
            return ImportRecordResult(
                row_number=row_number,
                status=ImportStatus.ERROR,
                module="dhg",
                dimension_type="subject",
                dimension_code=subject_code,
                value=None,
                message=f"Invalid FTE count: {fte_str}",
            )

        return ImportRecordResult(
            row_number=row_number,
            status=ImportStatus.VALID,
            module="dhg",
            dimension_type="subject",
            dimension_code=subject_code,
            value=fte,
        )

    def _validate_financial_record(
        self,
        row: dict[str, str],
        row_number: int,
        fiscal_year: int,
        module: ImportModule,
    ) -> ImportRecordResult:
        """Validate financial (revenue/costs) record."""
        account_code = self._get_column_value(row, FINANCIAL_COLUMNS["account_code"])
        amount_str = self._get_column_value(row, FINANCIAL_COLUMNS["annual_amount"])

        if not account_code:
            return ImportRecordResult(
                row_number=row_number,
                status=ImportStatus.ERROR,
                module=module.value,
                dimension_type="account_code",
                dimension_code=None,
                value=None,
                message="Missing account_code",
            )

        try:
            amount = Decimal(amount_str.replace(",", "")) if amount_str else Decimal("0")
        except (ValueError, TypeError, InvalidOperation):
            return ImportRecordResult(
                row_number=row_number,
                status=ImportStatus.ERROR,
                module=module.value,
                dimension_type="account_code",
                dimension_code=account_code,
                value=None,
                message=f"Invalid amount: {amount_str}",
            )

        # Warn if account code doesn't match expected pattern
        status = ImportStatus.VALID
        message = None

        if module == ImportModule.REVENUE and not account_code.startswith("7"):
            status = ImportStatus.WARNING
            message = f"Account code {account_code} doesn't look like revenue (expected 7xxxx)"
        elif module == ImportModule.COSTS and not account_code.startswith("6"):
            status = ImportStatus.WARNING
            message = f"Account code {account_code} doesn't look like costs (expected 6xxxx)"

        return ImportRecordResult(
            row_number=row_number,
            status=status,
            module=module.value,
            dimension_type="account_code",
            dimension_code=account_code,
            value=amount,
            message=message,
        )

    def _create_record(
        self,
        row: dict[str, str],
        fiscal_year: int,
        module: ImportModule,
    ) -> HistoricalActuals | None:
        """Create HistoricalActuals record from row data."""
        if module == ImportModule.ENROLLMENT:
            level_code = self._get_column_value(row, ENROLLMENT_COLUMNS["level_code"])
            count_str = self._get_column_value(row, ENROLLMENT_COLUMNS["student_count"])

            if not level_code or not count_str:
                return None

            level_name = self._get_column_value(row, ENROLLMENT_COLUMNS["level_name"])
            return HistoricalActuals(
                fiscal_year=fiscal_year,
                module_code=module.value,
                dimension_type="level",
                dimension_code=level_code,
                dimension_name=level_name,
                annual_count=int(count_str),
                data_source="manual_upload",
            )

        elif module == ImportModule.DHG:
            subject_code = self._get_column_value(row, DHG_COLUMNS["subject_code"])
            subject_name = self._get_column_value(row, DHG_COLUMNS["subject_name"])
            fte_str = self._get_column_value(row, DHG_COLUMNS["fte_count"])
            hours_str = self._get_column_value(row, DHG_COLUMNS["hours"])

            if not subject_code or not fte_str:
                return None

            return HistoricalActuals(
                fiscal_year=fiscal_year,
                module_code=module.value,
                dimension_type="subject",
                dimension_code=subject_code,
                dimension_name=subject_name,
                annual_fte=Decimal(fte_str),
                annual_hours=Decimal(hours_str) if hours_str else None,
                data_source="manual_upload",
            )

        else:  # Revenue or Costs
            account_code = self._get_column_value(row, FINANCIAL_COLUMNS["account_code"])
            account_name = self._get_column_value(row, FINANCIAL_COLUMNS["account_name"])
            amount_str = self._get_column_value(row, FINANCIAL_COLUMNS["annual_amount"])

            if not account_code or not amount_str:
                return None

            return HistoricalActuals(
                fiscal_year=fiscal_year,
                module_code=module.value,
                dimension_type="account_code",
                dimension_code=account_code,
                dimension_name=account_name,
                annual_amount_sar=Decimal(amount_str.replace(",", "")),
                data_source="manual_upload",
            )

    async def _delete_existing(
        self,
        fiscal_year: int,
        module: ImportModule | None,
    ) -> int:
        """Delete existing historical data for a year/module."""
        stmt = delete(HistoricalActuals).where(
            HistoricalActuals.fiscal_year == fiscal_year
        )

        if module:
            stmt = stmt.where(HistoricalActuals.module_code == module.value)

        result = await self.session.execute(stmt)
        return result.rowcount or 0


def generate_template(module: ImportModule) -> bytes:
    """
    Generate an Excel template for a module.

    Args:
        module: Module type

    Returns:
        Excel file bytes
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = f"{module.value}_template"

    # Set headers based on module
    if module == ImportModule.ENROLLMENT:
        headers = ["fiscal_year", "level_code", "level_name", "student_count"]
        example = [2024, "6EME", "Sixième", 115]
    elif module == ImportModule.DHG:
        headers = ["fiscal_year", "subject_code", "subject_name", "fte_count", "hours"]
        example = [2024, "MATH", "Mathematics", 5.5, 396]
    else:  # Revenue or Costs
        headers = ["fiscal_year", "account_code", "account_name", "annual_amount_sar"]
        if module == ImportModule.REVENUE:
            example = [2024, "70100", "Tuition Revenue", 45000000]
        else:
            example = [2024, "64100", "Teacher Salaries", 28000000]

    # Write headers
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    # Write example row
    for col, value in enumerate(example, start=1):
        ws.cell(row=2, column=col, value=value)

    # Set column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 20

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
