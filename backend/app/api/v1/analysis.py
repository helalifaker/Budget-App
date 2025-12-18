"""
Analysis API endpoints.

Provides REST API for:
- KPI calculations and benchmarking
- Dashboard summaries and charts
- Budget vs actual variance analysis
- Strategic planning and scenario modeling
"""

import uuid

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.dependencies.auth import UserDep
from app.schemas.insights import (
    ActivityLogEntry,
    AlertResponse,
    ChartDataResponse,
    ComparisonResponse,
    DashboardSummaryResponse,
    ForecastRevisionRequest,
    ForecastRevisionResponse,
    KPIBenchmarkComparison,
    KPICalculationRequest,
    KPITrendResponse,
    KPIValueResponse,
    MessageResponse,
    ScenarioComparisonResponse,
    StrategicInitiativeCreate,
    StrategicInitiativeResponse,
    StrategicPlanCreate,
    StrategicPlanResponse,
    UpdateAssumptionsRequest,
    VarianceReportResponse,
    YearProjectionResponse,
)
from app.services.admin.materialized_view_service import MaterializedViewService
from app.services.admin.strategic_service import StrategicService
from app.services.exceptions import (
    BusinessRuleError,
    NotFoundError,
    ValidationError,
)
from app.services.insights.budget_actual_service import BudgetActualService
from app.services.insights.dashboard_service import DashboardService
from app.services.insights.kpi_service import KPIService

router = APIRouter(prefix="/analysis", tags=["analysis"])


# ============================================================================
# Service Dependencies
# ============================================================================


def get_kpi_service(db: AsyncSession = Depends(get_db)) -> KPIService:
    """Get KPI service instance."""
    return KPIService(db)


def get_dashboard_service(db: AsyncSession = Depends(get_db)) -> DashboardService:
    """Get dashboard service instance."""
    return DashboardService(db)


def get_budget_actual_service(
    db: AsyncSession = Depends(get_db),
) -> BudgetActualService:
    """Get budget actual service instance."""
    return BudgetActualService(db)


def get_strategic_service(db: AsyncSession = Depends(get_db)) -> StrategicService:
    """Get strategic service instance."""
    return StrategicService(db)


# ============================================================================
# KPI Endpoints
# ============================================================================


@router.post(
    "/kpis/{version_id}/calculate",
    response_model=MessageResponse,
    status_code=status.HTTP_200_OK,
)
async def calculate_kpis(
    version_id: uuid.UUID,
    request: KPICalculationRequest | None = None,
    current_user: UserDep = None,
    kpi_service: KPIService = Depends(get_kpi_service),
    db: AsyncSession = Depends(get_db),
):
    """
    Calculate KPIs for a budget version.

    Calculates all active KPIs or specific KPIs and saves to database.
    """
    try:
        kpi_codes = request.kpi_codes if request else None
        results = await kpi_service.calculate_kpis(version_id, kpi_codes)
        saved_values = await kpi_service.save_kpi_values(version_id, results)
        await db.commit()

        return MessageResponse(
            message=f"Successfully calculated and saved {len(saved_values)} KPIs",
            details={
                "version_id": str(version_id),
                "kpi_count": len(saved_values),
                "kpi_codes": [v.kpi_definition.code for v in saved_values],
            },
        )
    except NotFoundError as e:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(e))
    except ValidationError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.get(
    "/kpis/{version_id}",
    response_model=list[KPIValueResponse],
)
async def get_all_kpis(
    version_id: uuid.UUID,
    category: str | None = Query(None, description="Filter by category"),
    current_user: UserDep = None,
    kpi_service: KPIService = Depends(get_kpi_service),
):
    """
    Get all calculated KPIs for a budget version.

    Optionally filter by category (educational, financial, operational, strategic).
    """
    try:
        from app.models import KPICategory

        category_filter = KPICategory(category) if category else None
        kpi_values = await kpi_service.get_all_kpis(version_id, category_filter)

        return [
            KPIValueResponse(
                id=kv.id,
                version_id=kv.version_id,
                kpi_code=kv.kpi_definition.code,
                kpi_name=kv.kpi_definition.name_en,
                calculated_value=kv.calculated_value,
                variance_from_target=kv.variance_from_target,
                variance_percent=kv.variance_percent,
                calculation_inputs=kv.calculation_inputs,
                calculated_at=kv.calculated_at,
                unit=kv.kpi_definition.unit,
                target_value=kv.kpi_definition.target_value,
                notes=kv.notes,
            )
            for kv in kpi_values
        ]
    except NotFoundError as e:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(e))
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.get(
    "/kpis/{version_id}/{kpi_code}",
    response_model=KPIValueResponse,
)
async def get_kpi_by_type(
    version_id: uuid.UUID,
    kpi_code: str,
    current_user: UserDep = None,
    kpi_service: KPIService = Depends(get_kpi_service),
):
    """Get specific KPI value for a budget version."""
    try:
        kpi_value = await kpi_service.get_kpi_by_type(version_id, kpi_code)

        if not kpi_value:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"KPI '{kpi_code}' not found for version {version_id}",
            )

        return KPIValueResponse(
            id=kpi_value.id,
            version_id=kpi_value.version_id,
            kpi_code=kpi_value.kpi_definition.code,
            kpi_name=kpi_value.kpi_definition.name_en,
            calculated_value=kpi_value.calculated_value,
            variance_from_target=kpi_value.variance_from_target,
            variance_percent=kpi_value.variance_percent,
            calculation_inputs=kpi_value.calculation_inputs,
            calculated_at=kpi_value.calculated_at,
            unit=kpi_value.kpi_definition.unit,
            target_value=kpi_value.kpi_definition.target_value,
            notes=kpi_value.notes,
        )
    except NotFoundError as e:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(e))


@router.get(
    "/kpis/trends/{kpi_code}",
    response_model=KPITrendResponse,
)
async def get_kpi_trends(
    kpi_code: str,
    version_ids: list[uuid.UUID] = Query(..., description="List of version IDs"),
    current_user: UserDep = None,
    kpi_service: KPIService = Depends(get_kpi_service),
):
    """Get KPI trend across multiple budget versions."""
    try:
        definition = await kpi_service.get_kpi_definition(kpi_code)
        trends = await kpi_service.get_kpi_trends(version_ids, kpi_code)

        from app.schemas.analysis import KPITrendPoint

        return KPITrendResponse(
            kpi_code=kpi_code,
            kpi_name=definition.name_en,
            unit=definition.unit,
            trend_points=[KPITrendPoint(**point) for point in trends],
        )
    except NotFoundError as e:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(e))


@router.get(
    "/kpis/{version_id}/benchmarks",
    response_model=dict[str, KPIBenchmarkComparison],
)
async def get_benchmark_comparison(
    version_id: uuid.UUID,
    current_user: UserDep = None,
    kpi_service: KPIService = Depends(get_kpi_service),
):
    """Compare budget version KPIs to AEFE benchmarks."""
    try:
        comparison = await kpi_service.get_benchmark_comparison(version_id)
        return {
            code: KPIBenchmarkComparison(**data) for code, data in comparison.items()
        }
    except NotFoundError as e:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(e))


# ============================================================================
# Dashboard Endpoints
# ============================================================================


@router.get(
    "/dashboard/{version_id}/summary",
    response_model=DashboardSummaryResponse,
)
async def get_dashboard_summary(
    version_id: uuid.UUID,
    current_user: UserDep = None,
    dashboard_service: DashboardService = Depends(get_dashboard_service),
):
    """Get dashboard summary cards for a budget version."""
    try:
        summary = await dashboard_service.get_dashboard_summary(version_id)
        return DashboardSummaryResponse(**summary)
    except NotFoundError as e:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(e))


@router.get(
    "/dashboard/{version_id}/charts/enrollment",
    response_model=ChartDataResponse,
)
async def get_enrollment_chart(
    version_id: uuid.UUID,
    breakdown_by: str = Query("level", description="level, nationality, or cycle"),
    current_user: UserDep = None,
    dashboard_service: DashboardService = Depends(get_dashboard_service),
):
    """Get enrollment chart data."""
    try:
        chart_data = await dashboard_service.get_enrollment_chart_data(
            version_id, breakdown_by
        )
        return ChartDataResponse(**chart_data)
    except NotFoundError as e:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(e))


@router.get(
    "/dashboard/{version_id}/charts/costs",
    response_model=ChartDataResponse,
)
async def get_cost_breakdown_chart(
    version_id: uuid.UUID,
    breakdown_by: str = Query("category", description="category, account, or period"),
    current_user: UserDep = None,
    dashboard_service: DashboardService = Depends(get_dashboard_service),
):
    """Get cost breakdown chart data."""
    try:
        chart_data = await dashboard_service.get_cost_breakdown(version_id, breakdown_by)
        return ChartDataResponse(**chart_data)
    except NotFoundError as e:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(e))


@router.get(
    "/dashboard/{version_id}/charts/revenue",
    response_model=ChartDataResponse,
)
async def get_revenue_breakdown_chart(
    version_id: uuid.UUID,
    breakdown_by: str = Query(
        "fee_type", description="fee_type, nationality, trimester, or period"
    ),
    current_user: UserDep = None,
    dashboard_service: DashboardService = Depends(get_dashboard_service),
):
    """Get revenue breakdown chart data."""
    try:
        chart_data = await dashboard_service.get_revenue_breakdown(
            version_id, breakdown_by
        )
        return ChartDataResponse(**chart_data)
    except NotFoundError as e:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(e))


@router.get(
    "/dashboard/{version_id}/alerts",
    response_model=list[AlertResponse],
)
async def get_alerts(
    version_id: uuid.UUID,
    current_user: UserDep = None,
    dashboard_service: DashboardService = Depends(get_dashboard_service),
):
    """Get system alerts for a budget version."""
    try:
        alerts = await dashboard_service.get_alerts(version_id)
        return [AlertResponse(**alert) for alert in alerts]
    except NotFoundError as e:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(e))


@router.get(
    "/dashboard/{version_id}/activity",
    response_model=list[ActivityLogEntry],
)
async def get_recent_activity(
    version_id: uuid.UUID | None = None,
    limit: int = Query(20, ge=1, le=100, description="Maximum activities to return"),
    current_user: UserDep = None,
    dashboard_service: DashboardService = Depends(get_dashboard_service),
):
    """Get recent activity log."""
    activities = await dashboard_service.get_recent_activity(version_id, limit)
    return [ActivityLogEntry(**activity) for activity in activities]


@router.get(
    "/activity",
    response_model=list[ActivityLogEntry],
)
async def get_recent_activity_no_version(
    limit: int = Query(10, ge=1, le=100, description="Maximum activities to return"),
    current_user: UserDep = None,
    dashboard_service: DashboardService = Depends(get_dashboard_service),
):
    """Get recent activity log (without version_id)."""
    activities = await dashboard_service.get_recent_activity(None, limit)
    return [ActivityLogEntry(**activity) for activity in activities]


@router.get(
    "/dashboard/compare",
    response_model=ComparisonResponse,
)
async def get_comparison_data(
    version_ids: list[uuid.UUID] = Query(..., description="List of version IDs"),
    metric: str = Query("summary", description="Metric to compare"),
    current_user: UserDep = None,
    dashboard_service: DashboardService = Depends(get_dashboard_service),
):
    """Get comparison data across multiple budget versions."""
    comparison = await dashboard_service.get_comparison_data(version_ids, metric)
    return ComparisonResponse(**comparison)


# ============================================================================
# Budget vs Actual Endpoints
# ============================================================================


@router.post(
    "/actuals/{version_id}/import",
    status_code=status.HTTP_201_CREATED,
)
async def import_actuals(
    version_id: uuid.UUID,
    file: UploadFile = File(..., description="Excel or CSV file with actual data"),
    period: str = Form(..., description="Period to import (T1, T2, T3, or ANNUAL)"),
    current_user: UserDep = None,
    budget_actual_service: BudgetActualService = Depends(get_budget_actual_service),
    db: AsyncSession = Depends(get_db),
):
    """
    Import actual financial data from Excel/CSV file.

    The file should contain columns:
    - account_code: PCG account code
    - description: Account description (optional)
    - amount_sar: Amount in SAR

    Args:
        version_id: Budget version UUID
        file: Excel (.xlsx) or CSV file with actual data
        period: Period to import (T1, T2, T3, or ANNUAL)
        current_user: Current authenticated user
        budget_actual_service: Budget actual service
        db: Database session

    Returns:
        Import result with count and status
    """
    import csv
    import io
    from datetime import datetime

    try:
        # Validate file type
        filename = file.filename or ""
        if not filename.endswith((".csv", ".xlsx", ".xls")):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="File must be CSV or Excel format (.csv, .xlsx, .xls)",
            )

        # Validate period
        valid_periods = ["T1", "T2", "T3", "ANNUAL"]
        if period.upper() not in valid_periods:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Period must be one of: {', '.join(valid_periods)}",
            )

        # Read file content
        content = await file.read()

        # Parse data based on file type
        records = []
        if filename.endswith(".csv"):
            # Parse CSV
            text = content.decode("utf-8")
            reader = csv.DictReader(io.StringIO(text))
            for row in reader:
                record = {
                    "account_code": row.get("account_code", ""),
                    "description": row.get("description", ""),
                    "amount_sar": float(row.get("amount_sar", 0) or 0),
                    "period": period.upper(),
                }
                if record["account_code"]:
                    records.append(record)
        else:
            # Parse Excel
            try:
                import openpyxl

                wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
                ws = wb.active
                if ws is None:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="Excel file has no active worksheet",
                    )

                # Get headers from first row
                headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                headers = [str(h).lower().strip() if h else "" for h in headers]

                # Map common header variations
                account_col = next(
                    (i for i, h in enumerate(headers) if h in ["account_code", "account", "code"]),
                    None,
                )
                desc_col = next(
                    (i for i, h in enumerate(headers) if h in ["description", "desc", "name"]),
                    None,
                )
                amount_col = next(
                    (i for i, h in enumerate(headers) if h in ["amount_sar", "amount", "value"]),
                    None,
                )

                if account_col is None or amount_col is None:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="Excel file must have 'account_code' and 'amount_sar' columns",
                    )

                for row in ws.iter_rows(min_row=2, values_only=True):
                    account_code = str(row[account_col] or "") if account_col < len(row) else ""
                    description = str(row[desc_col] or "") if desc_col is not None and desc_col < len(row) else ""
                    try:
                        amount = float(row[amount_col] or 0) if amount_col < len(row) else 0
                    except (ValueError, TypeError):
                        amount = 0

                    if account_code:
                        records.append({
                            "account_code": account_code,
                            "description": description,
                            "amount_sar": amount,
                            "period": period.upper(),
                        })

                wb.close()
            except ImportError:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Excel file support requires openpyxl. Please use CSV format.",
                )

        if not records:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No valid records found in file",
            )

        # Import records using the service
        result = await budget_actual_service.import_actuals_from_file(
            version_id=version_id,
            records=records,
            period=period.upper(),
            user_id=current_user.user_id if current_user else None,
        )
        await db.commit()

        return {
            "success": True,
            "imported_count": result.get("records_imported", len(records)),
            "period": period.upper(),
            "import_date": datetime.utcnow().isoformat(),
        }
    except HTTPException:
        raise
    except NotFoundError as e:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(e))
    except ValidationError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to import file: {e!s}",
        )


@router.post(
    "/actuals/{version_id}/calculate-variance",
    response_model=MessageResponse,
)
async def calculate_variance(
    version_id: uuid.UUID,
    period: int = Query(..., ge=1, le=12, description="Month (1-12)"),
    account_code: str | None = Query(None, description="Optional specific account"),
    current_user: UserDep = None,
    budget_actual_service: BudgetActualService = Depends(get_budget_actual_service),
    db: AsyncSession = Depends(get_db),
):
    """Calculate budget vs actual variance for a period."""
    try:
        variances = await budget_actual_service.calculate_variance(
            version_id, period, account_code
        )
        await db.commit()

        return MessageResponse(
            message=f"Successfully calculated {len(variances)} variances",
            details={
                "version_id": str(version_id),
                "period": period,
                "variance_count": len(variances),
            },
        )
    except NotFoundError as e:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(e))


@router.get(
    "/actuals/{version_id}/variance",
    response_model=VarianceReportResponse,
)
async def get_variance_report(
    version_id: uuid.UUID,
    period: int | None = Query(None, ge=1, le=12, description="Optional period filter"),
    account_code_pattern: str | None = Query(
        None, description="Account code pattern (e.g., '64%')"
    ),
    material_only: bool = Query(False, description="Only material variances"),
    current_user: UserDep = None,
    budget_actual_service: BudgetActualService = Depends(get_budget_actual_service),
):
    """Get comprehensive variance report."""
    try:
        report = await budget_actual_service.get_variance_report(
            version_id, period, account_code_pattern, material_only
        )
        return VarianceReportResponse(**report)
    except NotFoundError as e:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(e))


@router.post(
    "/actuals/{version_id}/forecast",
    response_model=ForecastRevisionResponse,
    status_code=status.HTTP_201_CREATED,
)
async def create_forecast_revision(
    version_id: uuid.UUID,
    request: ForecastRevisionRequest,
    current_user: UserDep = None,
    budget_actual_service: BudgetActualService = Depends(get_budget_actual_service),
    db: AsyncSession = Depends(get_db),
):
    """Create forecast revision based on actuals to date."""
    try:
        forecast = await budget_actual_service.create_forecast_revision(
            version_id, request.forecast_name, request.current_period
        )
        await db.commit()

        return ForecastRevisionResponse(
            forecast_version_id=str(forecast.id),
            forecast_name=forecast.name,
            fiscal_year=forecast.fiscal_year,
            status=forecast.status.value,
            created_at=forecast.created_at.isoformat(),
        )
    except NotFoundError as e:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(e))
    except BusinessRuleError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


# ============================================================================
# Strategic Planning Endpoints
# ============================================================================


@router.post(
    "/strategic-plans",
    response_model=StrategicPlanResponse,
    status_code=status.HTTP_201_CREATED,
)
async def create_strategic_plan(
    request: StrategicPlanCreate,
    current_user: UserDep = None,
    strategic_service: StrategicService = Depends(get_strategic_service),
    db: AsyncSession = Depends(get_db),
):
    """Create a new 5-year strategic plan."""
    try:
        plan = await strategic_service.create_strategic_plan(
            base_version_id=request.base_version_id,
            plan_name=request.plan_name,
            description=request.description,
            years=request.years,
            create_default_scenarios=request.create_default_scenarios,
        )
        await db.commit()
        await db.refresh(plan)

        return StrategicPlanResponse.model_validate(plan)
    except NotFoundError as e:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(e))
    except BusinessRuleError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    except ValidationError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.get(
    "/strategic-plans/{plan_id}",
    response_model=StrategicPlanResponse,
)
async def get_strategic_plan(
    plan_id: uuid.UUID,
    current_user: UserDep = None,
    strategic_service: StrategicService = Depends(get_strategic_service),
):
    """Get strategic plan by ID."""
    try:
        plan = await strategic_service.get_strategic_plan(plan_id)
        return StrategicPlanResponse.model_validate(plan)
    except NotFoundError as e:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(e))


@router.get(
    "/strategic-plans/{plan_id}/year/{year}",
    response_model=YearProjectionResponse,
)
async def get_year_projection(
    plan_id: uuid.UUID,
    year: int,
    scenario_type: str | None = Query(None, description="Optional scenario filter"),
    current_user: UserDep = None,
    strategic_service: StrategicService = Depends(get_strategic_service),
):
    """Get financial projections for a specific year."""
    try:
        from app.models import ScenarioType

        scenario_filter = ScenarioType(scenario_type) if scenario_type else None
        projections = await strategic_service.get_year_projections(
            plan_id, year, scenario_filter
        )
        return YearProjectionResponse(**projections)
    except NotFoundError as e:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(e))
    except ValidationError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.put(
    "/strategic-plans/scenarios/{scenario_id}/assumptions",
    response_model=MessageResponse,
)
async def update_scenario_assumptions(
    scenario_id: uuid.UUID,
    request: UpdateAssumptionsRequest,
    current_user: UserDep = None,
    strategic_service: StrategicService = Depends(get_strategic_service),
    db: AsyncSession = Depends(get_db),
):
    """Update scenario growth assumptions."""
    try:
        scenario = await strategic_service.update_assumptions(
            scenario_id=scenario_id,
            enrollment_growth_rate=request.enrollment_growth_rate,
            fee_increase_rate=request.fee_increase_rate,
            salary_inflation_rate=request.salary_inflation_rate,
            operating_inflation_rate=request.operating_inflation_rate,
            recalculate_projections=request.recalculate_projections,
        )
        await db.commit()

        return MessageResponse(
            message="Scenario assumptions updated successfully",
            details={
                "scenario_id": str(scenario.id),
                "scenario_type": scenario.scenario_type.value,
                "projections_recalculated": request.recalculate_projections,
            },
        )
    except NotFoundError as e:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(e))


@router.get(
    "/strategic-plans/{plan_id}/scenarios",
    response_model=ScenarioComparisonResponse,
)
async def compare_scenarios(
    plan_id: uuid.UUID,
    current_user: UserDep = None,
    strategic_service: StrategicService = Depends(get_strategic_service),
):
    """Compare all scenarios in a strategic plan."""
    try:
        comparison = await strategic_service.compare_scenarios(plan_id)
        return ScenarioComparisonResponse(**comparison)
    except NotFoundError as e:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(e))


@router.post(
    "/strategic-plans/{plan_id}/initiatives",
    response_model=StrategicInitiativeResponse,
    status_code=status.HTTP_201_CREATED,
)
async def add_initiative(
    plan_id: uuid.UUID,
    request: StrategicInitiativeCreate,
    current_user: UserDep = None,
    strategic_service: StrategicService = Depends(get_strategic_service),
    db: AsyncSession = Depends(get_db),
):
    """Add strategic initiative to a plan."""
    try:
        from app.models import InitiativeStatus

        initiative = await strategic_service.add_initiative(
            plan_id=plan_id,
            name=request.name,
            description=request.description,
            planned_year=request.planned_year,
            capex_amount_sar=request.capex_amount_sar,
            operating_impact_sar=request.operating_impact_sar,
            status=InitiativeStatus.PLANNED,
        )
        await db.commit()

        return StrategicInitiativeResponse.model_validate(initiative)
    except NotFoundError as e:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(e))
    except ValidationError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


# ============================================================================
# Materialized View Management Endpoints
# ============================================================================


@router.post(
    "/materialized-views/refresh-all",
    response_model=dict,
    status_code=status.HTTP_200_OK,
)
async def refresh_all_materialized_views(
    current_user: UserDep = None,
    db: AsyncSession = Depends(get_db),
):
    """
    Refresh all materialized views.

    This endpoint refreshes all KPI dashboard materialized views
    to ensure the latest data is available for queries.

    Use this endpoint:
    - After bulk data imports
    - After budget consolidation
    - On a scheduled basis (e.g., nightly)

    Performance: ~1-3 seconds total for all views.
    """
    results = await MaterializedViewService.refresh_all(db)

    # Count successes and failures
    success_count = sum(1 for r in results.values() if r["status"] == "success")
    error_count = sum(1 for r in results.values() if r["status"] == "error")

    return {
        "status": "completed",
        "total_views": len(results),
        "successful_refreshes": success_count,
        "failed_refreshes": error_count,
        "results": results,
    }


@router.post(
    "/materialized-views/refresh/{view_name}",
    response_model=dict,
    status_code=status.HTTP_200_OK,
)
async def refresh_specific_materialized_view(
    view_name: str,
    current_user: UserDep = None,
    db: AsyncSession = Depends(get_db),
):
    """
    Refresh a specific materialized view.

    Args:
        view_name: Name of the view to refresh (without schema prefix)
            - "mv_kpi_dashboard"
            - "mv_budget_consolidation"

    Performance: ~0.5-2 seconds depending on data volume.
    """
    # Add schema prefix if not present
    full_view_name = (
        f"efir_budget.{view_name}" if "." not in view_name else view_name
    )

    try:
        result = await MaterializedViewService.refresh_view(db, full_view_name)
        return result
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.get(
    "/materialized-views/info/{view_name}",
    response_model=dict,
    status_code=status.HTTP_200_OK,
)
async def get_materialized_view_info(
    view_name: str,
    current_user: UserDep = None,
    db: AsyncSession = Depends(get_db),
):
    """
    Get information about a materialized view.

    Returns row count, size, and other metadata.

    Args:
        view_name: Name of the view (without schema prefix)
            - "mv_kpi_dashboard"
            - "mv_budget_consolidation"
    """
    # Add schema prefix if not present
    full_view_name = (
        f"efir_budget.{view_name}" if "." not in view_name else view_name
    )

    try:
        info = await MaterializedViewService.get_view_info(db, full_view_name)
        return info
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
